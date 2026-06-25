"""Ki67 新版預測流程。"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Sequence

import numpy as np
import pandas as pd

from ki67_pred_utils import (
    DEFAULT_PREDICT_OUTPUT_DIR,
    DEFAULT_RESULTS_DIR,
    DEFAULT_TRAIN_OUTPUT_DIR,
    aggregate_to_image_features,
    attach_image_embeddings,
    build_feature_groups,
    evaluate_cell_predictions,
    evaluate_image_predictions,
    extract_image_embeddings,
    find_cleaned_csv_files,
    list_input_pc_image_stems,
    load_ki67_dataset,
    load_model_bundle,
    predict_feature_group_scores,
    predict_positive_probability,
    predict_stage2_ratio,
    reset_directory,
    safe_score_name,
    transform_table_features,
)

MODEL_BUNDLE_PATH = DEFAULT_TRAIN_OUTPUT_DIR / "model" / "ki67_model_bundle.joblib"
PREDICTION_SUMMARY_SHEET = "Prediction Summary"
PRED_RATIO_WEIGHTING_NOTE = (
    "pred_ratio 與 pred_ratio_weighted 皆為依 image cell_count 加權的 image-level "
    "pred_ratio 平均；pred_ratio_unweighted 為每張 image 等權平均。"
)
WORKBOOK_RATIO_WEIGHTING_NOTE = (
    "Prediction Summary Metrics 中不帶 weighted 後綴的 ratio/error 為每張 image 等權平均；"
    "*_weighted 為依 image cell_count 加權平均。"
)
FEATURE_GROUP_SHEET_NAMES = {
    "Texture": "texture",
    "Intensity distribution": "intensity",
    "Attachment / spreading": "attachment",
    "Halo / rounding": "halo",
    "Local crowding": "local_crowding",
    "Colony / FOV context": "colony_context",
    "Mitosis likelihood": "mitosis",
    "Nuclear morphology": "nuclear",
    "Debris / culture health": "debris",
}


def parse_args() -> argparse.Namespace:
    """解析命令列參數。"""
    parser = argparse.ArgumentParser(description="Predict Ki67 positive ratio with the new pipeline.")
    parser.add_argument("--results-dir", type=Path, default=DEFAULT_RESULTS_DIR)
    parser.add_argument("--model-bundle", type=Path, default=MODEL_BUNDLE_PATH)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_PREDICT_OUTPUT_DIR)
    parser.add_argument("--source-folder", action="append", default=None)
    parser.add_argument("--keep-output", action="store_true")
    parser.add_argument("--cnn-pretrained", action="store_true", default=None)
    parser.add_argument("--save-detail-table", action="store_true")
    return parser.parse_args()


def build_stage1_matrix_for_prediction(
    frame: pd.DataFrame,
    config: dict[str, Any],
    evidence_scores: pd.DataFrame,
    image_embeddings: pd.DataFrame,
    bundle: dict[str, Any],
) -> pd.DataFrame:
    """依訓練 config 建立預測用 Stage 1 matrix。"""
    parts: list[pd.DataFrame] = []
    if config.get("include_feature_scores", False):
        groups = config.get("feature_score_groups")
        if groups:
            score_columns = [safe_score_name(group_name) for group_name in groups]
            evidence_part = evidence_scores.reindex(columns=score_columns)
        else:
            evidence_part = evidence_scores
        parts.append(evidence_part.reset_index(drop=True))
    if config.get("include_cnn_embedding", False):
        parts.append(attach_image_embeddings(frame, image_embeddings).reset_index(drop=True))
    if config.get("parameter_mode") == "selected":
        parts.append(transform_table_features(frame, bundle["selected_parameter_transformer"]).reset_index(drop=True))
    elif config.get("parameter_mode") == "all":
        parts.append(transform_table_features(frame, bundle["all_parameter_transformer"]).reset_index(drop=True))

    if not parts:
        raise ValueError("模型 bundle 的 Stage 1 設定沒有任何輸入欄位。")
    matrix = pd.concat(parts, axis=1)
    matrix = matrix.loc[:, ~matrix.columns.duplicated()].copy()
    return matrix.reindex(columns=bundle["stage1_feature_columns"], fill_value=0.0).fillna(0.0)


def safe_weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    """計算可容忍缺值的加權平均。

    Args:
        values: 要平均的數值欄位。
        weights: 對應的權重欄位。

    Returns:
        加權平均；若沒有可用資料則回傳 ``nan``。
    """
    numeric_values = pd.to_numeric(values, errors="coerce")
    numeric_weights = pd.to_numeric(weights, errors="coerce")
    valid = numeric_values.notna() & numeric_weights.notna() & (numeric_weights > 0)
    if not valid.any():
        return float("nan")
    return float(np.average(numeric_values[valid], weights=numeric_weights[valid]))


def add_folder_image_diagnostics(folder_df: pd.DataFrame, image_df: pd.DataFrame) -> pd.DataFrame:
    """加入 data/input PC 影像數量比對資訊。"""
    rows = []
    for _, row in folder_df.iterrows():
        source_folder = str(row["source_folder"])
        expected = set(list_input_pc_image_stems(source_folder))
        used = set(
            image_df.loc[image_df["source_folder"].astype(str) == source_folder, "image_name"].astype(str)
        )
        result = row.to_dict()
        result["input_image_count"] = len(expected)
        result["used_image_count"] = len(used)
        result["missing_image_count"] = len(expected - used)
        result["extra_image_count"] = len(used - expected)
        result["missing_images"] = ";".join(sorted(expected - used))
        result["extra_images"] = ";".join(sorted(used - expected))
        rows.append(result)
    return pd.DataFrame(rows)


def summarize_folder_predictions(image_df: pd.DataFrame) -> pd.DataFrame:
    """將 image-level prediction 彙整到 source folder。"""
    rows: list[dict[str, Any]] = []
    for source_folder, part in image_df.groupby("source_folder", sort=True):
        weights = pd.to_numeric(part["cell_count"], errors="coerce")
        pred_weighted = safe_weighted_mean(part["pred_ratio"], weights)
        pred_unweighted = float(pd.to_numeric(part["pred_ratio"], errors="coerce").mean())
        row = {
            "source_folder": source_folder,
            "image_count": int(part["image_key"].nunique()),
            "cell_count": int(part["cell_count"].sum()),
            "pred_ratio": pred_weighted,
            "pred_ratio_weighted": pred_weighted,
            "pred_ratio_unweighted": pred_unweighted,
            "pred_ratio_weighting_note": PRED_RATIO_WEIGHTING_NOTE,
        }
        if "true_ratio" in part.columns and part["true_ratio"].notna().any():
            true_weighted = safe_weighted_mean(part["true_ratio"], weights)
            true_unweighted = float(pd.to_numeric(part["true_ratio"], errors="coerce").mean())
            row["true_ratio"] = true_weighted
            row["true_ratio_weighted"] = true_weighted
            row["true_ratio_unweighted"] = true_unweighted
            row["abs_error"] = abs(pred_weighted - true_weighted)
            row["abs_error_weighted"] = row["abs_error"]
            row["abs_error_unweighted"] = abs(pred_unweighted - true_unweighted)
        rows.append(row)
    return pd.DataFrame(rows)


def prediction_workbook_feature_groups(
    feature_groups: dict[str, list[str]],
) -> list[tuple[str, list[str]]]:
    """取得正式預測 Excel 要輸出的單一 feature group。

    Args:
        feature_groups: 由 ``build_feature_groups()`` 建立的 feature group 對照。

    Returns:
        排除複合 group 後的 ``(group_name, columns)`` 清單。
    """
    groups: list[tuple[str, list[str]]] = []
    for group_name, columns in feature_groups.items():
        if not columns:
            continue
        if "+" in group_name or group_name == "All PC parameters":
            continue
        groups.append((group_name, list(columns)))
    return groups


def excel_sheet_name(group_name: str, used_names: set[str]) -> str:
    """將 feature group 名稱轉為合法且不重複的 Excel sheet 名稱。

    Args:
        group_name: Feature group 名稱。
        used_names: 已使用過的 sheet 名稱集合。

    Returns:
        可寫入 Excel 的 sheet 名稱。
    """
    preferred = FEATURE_GROUP_SHEET_NAMES.get(group_name, group_name.lower())
    safe_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", preferred).strip("' ")
    safe_name = safe_name[:31] or "sheet"
    candidate = safe_name
    suffix = 2
    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{safe_name[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def safe_path_component(value: str) -> str:
    """將資料夾或檔名前綴轉成 Windows 可用路徑片段。"""
    safe_value = re.sub(r'[<>:"/\\|?*]+', "_", str(value)).strip(" .")
    return safe_value or "prediction"


def existing_ordered_columns(frame: pd.DataFrame, columns: Sequence[str]) -> list[str]:
    """依指定順序回傳目前 DataFrame 中存在的欄位。"""
    return [column for column in columns if column in frame.columns]


def excel_cell_value(value: Any) -> Any:
    """將 pandas / numpy 值轉成 openpyxl 可寫入的純量。"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        return value.item()
    return value


def iter_excel_rows(frame: pd.DataFrame):
    """逐列產生可寫入 Excel 的資料列。"""
    yield list(frame.columns)
    for row in frame.itertuples(index=False, name=None):
        yield [excel_cell_value(value) for value in row]


def write_dataframe_sheet(workbook: Any, sheet_name: str, frame: pd.DataFrame) -> None:
    """將 DataFrame 寫入 workbook 的單一 sheet。

    Args:
        workbook: openpyxl workbook。
        sheet_name: Sheet 名稱。
        frame: 要輸出的資料表。
    """
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.freeze_panes = "A2"
    for row in iter_excel_rows(frame):
        worksheet.append(row)


def write_feature_sheet(
    workbook: Any,
    sheet_name: str,
    frame: pd.DataFrame,
    id_columns: Sequence[str],
    feature_columns: Sequence[str],
) -> None:
    """輸出單一 feature group 的明細 sheet。

    Args:
        workbook: openpyxl workbook。
        sheet_name: Sheet 名稱。
        frame: 來源資料表。
        id_columns: 識別欄位。
        feature_columns: Feature 欄位。
    """
    columns = existing_ordered_columns(frame, [*id_columns, *feature_columns])
    write_dataframe_sheet(workbook, sheet_name, frame.loc[:, columns])


def write_prediction_summary_sheet(
    workbook: Any,
    metrics_df: pd.DataFrame,
    predictions_df: pd.DataFrame,
    metric_notes: Sequence[str] | None = None,
) -> None:
    """輸出與既有範例相同結構的 Prediction Summary sheet。

    Args:
        workbook: openpyxl workbook。
        metrics_df: 上半部 Metrics 區塊。
        predictions_df: 下半部 Predictions 區塊。
        metric_notes: Metrics 表格下方的補充說明。
    """
    worksheet = workbook.create_sheet(title=PREDICTION_SUMMARY_SHEET)
    worksheet.append(["Metrics"])
    for row in iter_excel_rows(metrics_df):
        worksheet.append(row)
    for note in metric_notes or []:
        worksheet.append([note])
    worksheet.append(["Predictions"])
    for row in iter_excel_rows(predictions_df):
        worksheet.append(row)


def with_split_column(frame: pd.DataFrame) -> pd.DataFrame:
    """補上與範例 workbook 相容的 split 欄位。"""
    result = frame.copy()
    if "split" not in result.columns:
        result.insert(0, "split", result["source_folder"].astype(str))
    return result


def append_prediction_columns(cell_df: pd.DataFrame, threshold: float = 0.5) -> pd.DataFrame:
    """在 single-cell table 加入預測類別與正確性欄位。

    Args:
        cell_df: Cell-level prediction table。
        threshold: 將 ``cell_prob`` 轉為陽性類別的門檻。

    Returns:
        加入 ``cell_pred`` 與選用 ``cell_correct`` 的資料表。
    """
    result = cell_df.copy()
    result["cell_prob"] = pd.to_numeric(result["cell_prob"], errors="coerce")
    result["cell_pred"] = (result["cell_prob"] >= float(threshold)).astype(int)
    if "ki67_label" in result.columns:
        label = pd.to_numeric(result["ki67_label"], errors="coerce")
        result["cell_correct"] = np.where(label.notna(), result["cell_pred"].eq(label.astype("Int64")), np.nan)
    return result


def append_image_error_columns(image_df: pd.DataFrame) -> pd.DataFrame:
    """在 image-level table 加入可用的誤差欄位。

    Args:
        image_df: Image-level prediction table。

    Returns:
        若有 ground truth，加入 ``signed_error`` 與 ``abs_error`` 的資料表。
    """
    result = image_df.copy()
    result["pred_ratio"] = pd.to_numeric(result["pred_ratio"], errors="coerce")
    if "true_ratio" in result.columns:
        result["true_ratio"] = pd.to_numeric(result["true_ratio"], errors="coerce")
        result["signed_error"] = result["pred_ratio"] - result["true_ratio"]
        result["abs_error"] = result["signed_error"].abs()
    return result


def aggregate_cell_features_to_image(
    cell_df: pd.DataFrame,
    image_df: pd.DataFrame,
    feature_columns: Sequence[str],
) -> pd.DataFrame:
    """將 cell-level feature 聚合成 image-level feature sheet。

    Args:
        cell_df: Cell-level prediction table。
        image_df: Image-level prediction table。
        feature_columns: 要聚合的 feature 欄位。

    Returns:
        每張 image 一列，feature 以 mean / median 聚合。
    """
    columns = existing_ordered_columns(cell_df, feature_columns)
    id_columns = existing_ordered_columns(image_df, ["image_key", "image_name"])
    if "image_key" not in id_columns:
        image_id_columns = existing_ordered_columns(image_df, ["image_name"])
        return image_df.loc[:, image_id_columns].drop_duplicates()
    id_frame = image_df.loc[:, id_columns].drop_duplicates("image_key")
    if not columns:
        return id_frame.drop(columns=["image_key"], errors="ignore")

    numeric = cell_df.loc[:, ["image_key", *columns]].copy()
    for column in columns:
        numeric[column] = pd.to_numeric(numeric[column], errors="coerce")
    stats = numeric.groupby("image_key", as_index=True)[columns].agg(["mean", "median"])
    stats.columns = [f"{column}_{stat_name}" for column, stat_name in stats.columns]
    return id_frame.merge(stats.reset_index(), on="image_key", how="left").drop(columns=["image_key"], errors="ignore")


def grouped_with_overall(frame: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """產生 overall 與各 split 的資料切片。"""
    groups: list[tuple[str, pd.DataFrame]] = [("overall", frame)]
    if "split" in frame.columns:
        groups.extend((str(name), part) for name, part in frame.groupby("split", sort=True))
    return groups


def build_cell_metric_summary(cell_df: pd.DataFrame, threshold: float = 0.5) -> pd.DataFrame:
    """建立 single-cell workbook 的 Metrics 區塊。

    Args:
        cell_df: Cell-level prediction table。
        threshold: 將 ``cell_prob`` 轉為陽性類別的門檻。

    Returns:
        Metrics 資料表；有 ground truth 時輸出 accuracy，否則輸出預測摘要。
    """
    has_label = "ki67_label" in cell_df.columns
    prob = pd.to_numeric(cell_df["cell_prob"], errors="coerce")
    valid_prob = prob.notna()
    if not valid_prob.any():
        return pd.DataFrame()
    if has_label:
        y_true = pd.to_numeric(cell_df["ki67_label"], errors="coerce")
        valid = y_true.notna() & valid_prob
        if not valid.any():
            return pd.DataFrame()
        metrics = evaluate_cell_predictions(y_true[valid].astype(int), prob[valid], threshold=threshold)
        row = {
            "cell_num": int(valid.sum()),
            "positive_ratio": float(y_true[valid].mean()),
            "accuracy": metrics["accuracy"],
            "probability_mae": float(np.mean(np.abs(prob[valid].to_numpy() - y_true[valid].to_numpy()))),
            "tn": metrics["tn"],
            "fp": metrics["fp"],
            "fn": metrics["fn"],
            "tp": metrics["tp"],
        }
    else:
        row = {
            "cell_num": int(valid_prob.sum()),
            "mean_cell_prob": float(prob[valid_prob].mean()),
            "predicted_positive_ratio": float((prob[valid_prob] >= float(threshold)).mean()),
            "threshold": float(threshold),
        }
    return pd.DataFrame([row])


def build_image_metric_summary(image_df: pd.DataFrame) -> pd.DataFrame:
    """建立 image-level workbook 的 Metrics 區塊。

    Args:
        image_df: Image-level prediction table。

    Returns:
        Metrics 資料表，包含加權與未加權的 source-folder 預測比例。
    """
    columns = [
        "image_num",
        "cell_num",
        "pred_ratio",
        "pred_ratio_weighted",
        "true_ratio",
        "true_ratio_weighted",
        "image_mae",
        "image_bias",
        "image_rmse",
        "within_10pp",
        "abs_error",
        "abs_error_weighted",
        "ratio_accuracy",
        "ratio_accuracy_weighted",
    ]
    pred = pd.to_numeric(image_df["pred_ratio"], errors="coerce")
    weights = pd.to_numeric(image_df["cell_count"], errors="coerce")
    row: dict[str, Any] = {
        "image_num": int(image_df["image_key"].nunique()),
        "cell_num": int(weights.fillna(0).sum()),
        "pred_ratio": float(pred.mean()),
        "pred_ratio_weighted": safe_weighted_mean(pred, weights),
    }
    if "true_ratio" in image_df.columns:
        valid = image_df.copy()
        valid["true_ratio"] = pd.to_numeric(valid["true_ratio"], errors="coerce")
        valid["pred_ratio"] = pd.to_numeric(valid["pred_ratio"], errors="coerce")
        valid = valid.dropna(subset=["true_ratio", "pred_ratio"])
        if not valid.empty:
            true_ratio = pd.to_numeric(valid["true_ratio"], errors="coerce")
            valid_pred = pd.to_numeric(valid["pred_ratio"], errors="coerce")
            valid_weights = pd.to_numeric(valid["cell_count"], errors="coerce")
            metrics = evaluate_image_predictions(valid, valid["pred_ratio"])
            abs_error_weighted = abs(
                safe_weighted_mean(valid_pred, valid_weights)
                - safe_weighted_mean(true_ratio, valid_weights)
            )
            abs_error_unweighted = abs(float(valid_pred.mean()) - float(true_ratio.mean()))
            row.update(
                    {
                        "true_ratio": float(true_ratio.mean()),
                        "true_ratio_weighted": safe_weighted_mean(true_ratio, valid_weights),
                        "image_mae": metrics["image_mae"],
                        "image_bias": metrics["image_bias"],
                        "image_rmse": metrics["image_rmse"],
                        "within_10pp": metrics["within_10pp"],
                        "abs_error": abs_error_unweighted,
                        "abs_error_weighted": abs_error_weighted,
                        "ratio_accuracy": 1.0 - abs_error_unweighted,
                        "ratio_accuracy_weighted": 1.0 - abs_error_weighted,
                    }
            )
    return pd.DataFrame([row]).reindex(columns=columns)


def export_source_prediction_workbooks(
    output_dir: Path,
    cell_df: pd.DataFrame,
    image_df: pd.DataFrame,
    feature_groups: dict[str, list[str]],
    score_columns: Sequence[str],
    threshold: float = 0.5,
) -> list[Path]:
    """依 source folder 輸出 image 與 single-cell prediction Excel。

    Args:
        output_dir: 正式預測輸出根目錄。
        cell_df: Cell-level prediction table。
        image_df: Image-level prediction table。
        feature_groups: Feature group 對照表。
        score_columns: Feature-group evidence score 欄位。
        threshold: 將 ``cell_prob`` 轉為陽性類別的門檻。

    Returns:
        實際輸出的 workbook 路徑清單。
    """
    from openpyxl import Workbook

    output_paths: list[Path] = []
    workbook_groups = prediction_workbook_feature_groups(feature_groups)
    cell_df = append_prediction_columns(with_split_column(cell_df), threshold=threshold)
    image_df = append_image_error_columns(with_split_column(image_df))

    cell_feature_id_columns = ["Image", "Cell_ID"]
    cell_summary_columns = [
        "split",
        "source_folder",
        "Image",
        "Cell_ID",
        "image_key",
        "passage",
        "ki67_label",
        "cell_prob",
        "cell_pred",
        "cell_correct",
        *score_columns,
    ]
    image_summary_columns = [
        "split",
        "source_folder",
        "image_name",
        "image_key",
        "passage",
        "cell_count",
        "true_ratio",
        "pred_ratio",
        "signed_error",
        "abs_error",
        "mean_cell_prob",
        "median_cell_prob",
        "p90_cell_prob",
        "frac_prob_gt_0_3",
        "frac_prob_gt_0_5",
        "frac_prob_gt_0_7",
        *[
            score_column
            for column in score_columns
            for score_column in (f"{column}_mean", f"{column}_median")
        ],
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    for source_folder, source_cell in cell_df.groupby("source_folder", sort=True):
        source_name = str(source_folder)
        source_image = image_df[image_df["source_folder"].astype(str) == source_name].copy()
        source_output_dir = output_dir / safe_path_component(source_name)
        source_output_dir.mkdir(parents=True, exist_ok=True)
        file_prefix = f"ki67_{safe_path_component(source_name)}"

        cell_workbook = source_output_dir / f"{file_prefix}_single_cell_predictions.xlsx"
        cell_book = Workbook(write_only=True)
        used_names: set[str] = set()
        for group_name, columns in workbook_groups:
            write_feature_sheet(
                cell_book,
                excel_sheet_name(group_name, used_names),
                source_cell,
                cell_feature_id_columns,
                columns,
            )
        write_prediction_summary_sheet(
            cell_book,
            build_cell_metric_summary(source_cell, threshold=threshold),
            source_cell.loc[:, existing_ordered_columns(source_cell, cell_summary_columns)],
        )
        cell_book.save(cell_workbook)
        output_paths.append(cell_workbook)

        image_workbook = source_output_dir / f"{file_prefix}_image_predictions.xlsx"
        image_book = Workbook(write_only=True)
        used_names = set()
        image_feature_id_columns = ["image_name"]
        for group_name, columns in workbook_groups:
            image_feature_table = aggregate_cell_features_to_image(source_cell, source_image, columns)
            write_feature_sheet(
                image_book,
                excel_sheet_name(group_name, used_names),
                image_feature_table,
                image_feature_id_columns,
                [column for column in image_feature_table.columns if column not in set(image_feature_id_columns)],
            )
        write_prediction_summary_sheet(
            image_book,
            build_image_metric_summary(source_image),
            source_image.loc[:, existing_ordered_columns(source_image, image_summary_columns)],
            metric_notes=[f"ratio_weighting_note: {WORKBOOK_RATIO_WEIGHTING_NOTE}"],
        )
        image_book.save(image_workbook)
        output_paths.append(image_workbook)

    return output_paths


def run_prediction(args: argparse.Namespace) -> dict[str, pd.DataFrame]:
    """執行新版 Ki67 預測。

    Args:
        args: 命令列參數，需包含 results-dir、model-bundle 與 output-dir。

    Returns:
        dict[str, pd.DataFrame]: cell、image 與 folder 三層預測結果。
    """
    if not args.model_bundle.exists():
        raise FileNotFoundError(f"找不到模型 bundle: {args.model_bundle}")
    bundle = load_model_bundle(args.model_bundle)

    csv_files = find_cleaned_csv_files(
        results_dir=args.results_dir,
        include_source_folders=args.source_folder,
    )
    dataset = load_ki67_dataset(csv_files, require_label=False)
    if len(dataset) == 0:
        raise ValueError("沒有可預測的 cleaned CSV。")

    output_dir = args.output_dir
    if args.keep_output:
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        reset_directory(output_dir)

    evidence_scores = predict_feature_group_scores(dataset, bundle["feature_group_models"])
    config = dict(bundle["config"])
    if config.get("include_cnn_embedding", False):
        pretrained = bool(bundle.get("cnn_pretrained", False))
        if args.cnn_pretrained is not None:
            pretrained = bool(args.cnn_pretrained)
        image_embeddings, cnn_meta = extract_image_embeddings(
            dataset,
            pretrained=pretrained,
            random_state=int(bundle.get("random_state", 42)),
        )
        if image_embeddings.empty:
            raise RuntimeError(f"CNN embedding 為空，無法執行此模型。狀態: {cnn_meta}")
    else:
        image_embeddings = pd.DataFrame(index=[])

    x_pred = build_stage1_matrix_for_prediction(dataset, config, evidence_scores, image_embeddings, bundle)
    cell_prob = predict_positive_probability(bundle["stage1_model"], x_pred)
    cell_df = dataset.reset_index(drop=True).copy()
    for column in evidence_scores.columns:
        cell_df[column] = evidence_scores.reset_index(drop=True)[column].to_numpy(dtype=np.float64)
    cell_df["cell_prob"] = cell_prob
    cell_df["predicted_ki67_positive"] = (cell_df["cell_prob"] >= float(bundle.get("cell_threshold", 0.5))).astype(int)

    if "ki67_label" in cell_df.columns:
        cell_df["prediction_correct"] = cell_df["predicted_ki67_positive"] == cell_df["ki67_label"].astype(int)

    image_df = aggregate_to_image_features(cell_df, "cell_prob", list(evidence_scores.columns))
    image_df["pred_ratio"] = predict_stage2_ratio(
        image_df,
        bundle["stage2_model"],
        config["stage2_mode"],
        bundle["stage2_feature_columns"],
    )
    folder_df = summarize_folder_predictions(image_df)
    folder_df = add_folder_image_diagnostics(folder_df, image_df)
    workbook_paths = export_source_prediction_workbooks(
        output_dir=output_dir,
        cell_df=cell_df,
        image_df=image_df,
        feature_groups=build_feature_groups(dataset.columns),
        score_columns=list(evidence_scores.columns),
        threshold=float(bundle.get("cell_threshold", 0.5)),
    )

    cell_columns = [
        col
        for col in [
            "source_folder",
            "Image",
            "Cell_ID",
            "_source_row_id",
            "ki67_label",
            *list(evidence_scores.columns),
            "cell_prob",
            "predicted_ki67_positive",
            "prediction_correct",
        ]
        if col in cell_df.columns
    ]
    image_columns = [
        col
        for col in [
            "source_folder",
            "image_name",
            "image_key",
            "cell_count",
            "true_ratio",
            "pred_ratio",
            "mean_cell_prob",
            "median_cell_prob",
            "p90_cell_prob",
            "frac_prob_gt_0_3",
            "frac_prob_gt_0_5",
            "frac_prob_gt_0_7",
        ]
        if col in image_df.columns
    ]
    folder_df.to_csv(output_dir / "ki67_predictions.csv", index=False, encoding="utf-8-sig")

    if args.save_detail_table:
        folder_table = folder_df.copy()
        folder_table.insert(0, "level", "folder")
        image_table = image_df.loc[:, image_columns].copy()
        image_table.insert(0, "level", "image")
        cell_table = cell_df.loc[:, cell_columns].copy()
        cell_table.insert(0, "level", "cell")
        pd.concat([folder_table, image_table, cell_table], ignore_index=True, sort=False).to_csv(
            output_dir / "ki67_prediction_details.csv",
            index=False,
            encoding="utf-8-sig",
        )

    if "ki67_label" in cell_df.columns and "true_ratio" in image_df.columns:
        cell_metrics = evaluate_cell_predictions(cell_df["ki67_label"], cell_df["cell_prob"])
        image_metrics = evaluate_image_predictions(image_df.dropna(subset=["true_ratio"]), image_df.dropna(subset=["true_ratio"])["pred_ratio"])
        print(
            "評估結果 | "
            f"cell accuracy={100 * cell_metrics['accuracy']:.2f}% | "
            f"image MAE={100 * image_metrics['image_mae']:.2f} pp"
        )

    print(f"新版預測輸出: {output_dir}")
    print(f"Excel workbooks: {len(workbook_paths)}")
    return {"cell": cell_df, "image": image_df, "folder": folder_df}


def main() -> None:
    """命令列入口。"""
    run_prediction(parse_args())


if __name__ == "__main__":
    main()
