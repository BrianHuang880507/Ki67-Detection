"""Ki67 新版 Stage 1 / Stage 2 訓練流程。"""

from __future__ import annotations

import argparse
import re
from dataclasses import asdict
from pathlib import Path
from typing import Any, Sequence

import numpy as np
import pandas as pd

from ki67_pred_utils import (
    DEFAULT_EXCLUDED_SOURCE_FOLDERS,
    DEFAULT_RESULTS_DIR,
    DEFAULT_SUMMARY_HTML,
    DEFAULT_TRAIN_OUTPUT_DIR,
    ExperimentConfig,
    aggregate_to_image_features,
    attach_image_embeddings,
    build_feature_groups,
    default_experiment_configs,
    detect_numeric_feature_columns,
    evaluate_cell_predictions,
    evaluate_image_predictions,
    extract_image_embeddings,
    filter_extreme_image_ratios,
    find_cleaned_csv_files,
    fit_feature_group_models,
    fit_selected_parameter_transformer,
    fit_stage1_classifier_with_oof,
    fit_stage2_ratio_model,
    fit_table_transformer,
    format_percent,
    format_pp,
    load_ki67_dataset,
    predict_feature_group_scores,
    predict_positive_probability,
    predict_stage2_ratio,
    render_summary_html,
    reset_directory,
    safe_score_name,
    save_json,
    save_model_bundle,
    split_images_within_source_folder,
    summarize_split,
    transform_table_features,
)

pd.set_option("display.max_columns", 200)
pd.set_option("display.width", 220)

RANDOM_STATE = 42
BASE_FEATURE_GROUPS = ("Texture", "Intensity distribution", "Attachment / spreading", "Halo / rounding")
FORMAL_CONFIG_KEY = "texture_attachment_s1"
FEATURE_GROUP_SHEET_NAMES = {
    "Texture": "texture",
    "Intensity distribution": "intensity",
    "Attachment / spreading": "attachment",
    "Halo / rounding": "halo",
    "Local crowding": "local_crowding",
    "Colony / FOV context": "colony_context",
    "Mitosis likelihood": "mitosis",
    "Nuclear morphology": "nuclear",
    "Debris / culture health": "debris",
}
PREDICTION_SUMMARY_SHEET = "Prediction Summary"


def prediction_workbook_feature_groups(
    feature_groups: dict[str, list[str]],
) -> list[tuple[str, list[str]]]:
    """取得整理版 Excel 需要輸出的基礎特徵群。

    Args:
        feature_groups: 由 `build_feature_groups()` 建立的特徵群對照。

    Returns:
        只包含基礎特徵群的 `(群組名稱, 欄位列表)`。
    """
    groups: list[tuple[str, list[str]]] = []
    for group_name, columns in feature_groups.items():
        if not columns:
            continue
        if "+" in group_name or group_name == "All PC parameters":
            continue
        groups.append((group_name, list(columns)))
    return groups


def excel_sheet_name(group_name: str, used_names: set[str]) -> str:
    """將特徵群名稱轉成合法且不重複的 Excel 工作表名稱。

    Args:
        group_name: 特徵群名稱。
        used_names: 目前已使用的工作表名稱集合。

    Returns:
        合法的 Excel 工作表名稱。
    """
    preferred = FEATURE_GROUP_SHEET_NAMES.get(group_name, group_name.lower())
    safe_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", preferred).strip("' ")
    safe_name = safe_name[:31] or "sheet"
    candidate = safe_name
    suffix = 2
    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{safe_name[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def existing_ordered_columns(frame: pd.DataFrame, columns: Sequence[str]) -> list[str]:
    """依指定順序取出表格中實際存在的欄位。"""
    return [column for column in columns if column in frame.columns]


def feature_score_columns_for_config(config: ExperimentConfig, evidence_scores: pd.DataFrame) -> list[str]:
    """依模型 config 取得要使用的 feature-group evidence 欄位。

    Args:
        config: 實驗組合設定。
        evidence_scores: 已計算好的 feature-group evidence scores。

    Returns:
        Stage 1 要使用的 evidence score 欄位。
    """
    if config.feature_score_groups is None:
        return list(evidence_scores.columns)
    requested = [safe_score_name(group_name) for group_name in config.feature_score_groups]
    return [column for column in requested if column in evidence_scores.columns]


def append_prediction_columns(cell_df: pd.DataFrame, threshold: float = 0.5) -> pd.DataFrame:
    """替 single-cell prediction table 補上預測類別與是否正確。

    Args:
        cell_df: cell-level prediction table。
        threshold: 將 `cell_prob` 轉成陽性預測的門檻。

    Returns:
        已補上 `cell_pred`、`cell_correct` 的表格。
    """
    result = cell_df.copy()
    result["cell_prob"] = pd.to_numeric(result["cell_prob"], errors="coerce")
    result["cell_pred"] = (result["cell_prob"] >= float(threshold)).astype(int)
    if "ki67_label" in result.columns:
        label = pd.to_numeric(result["ki67_label"], errors="coerce")
        result["cell_correct"] = np.where(label.notna(), result["cell_pred"].eq(label.astype("Int64")), np.nan)
    return result


def append_image_error_columns(image_df: pd.DataFrame) -> pd.DataFrame:
    """替 image-level prediction table 補上誤差欄位。

    Args:
        image_df: image-level prediction table。

    Returns:
        已補上 `signed_error`、`abs_error` 的表格。
    """
    result = image_df.copy()
    result["true_ratio"] = pd.to_numeric(result["true_ratio"], errors="coerce")
    result["pred_ratio"] = pd.to_numeric(result["pred_ratio"], errors="coerce")
    result["signed_error"] = result["pred_ratio"] - result["true_ratio"]
    result["abs_error"] = result["signed_error"].abs()
    return result


def build_cell_metric_summary(cell_df: pd.DataFrame, threshold: float = 0.5) -> pd.DataFrame:
    """建立 single-cell 層級的整體與 split 指標。

    Args:
        cell_df: cell-level prediction table。
        threshold: 將 `cell_prob` 轉成陽性預測的門檻。

    Returns:
        指標摘要表。
    """
    rows: list[dict[str, Any]] = []
    for split_name, local in [("overall", cell_df), *list(cell_df.groupby("split", sort=True))]:
        if local.empty:
            continue
        y_true = pd.to_numeric(local["ki67_label"], errors="coerce")
        prob = pd.to_numeric(local["cell_prob"], errors="coerce")
        valid = y_true.notna() & prob.notna()
        if not valid.any():
            continue
        metrics = evaluate_cell_predictions(y_true[valid].astype(int), prob[valid], threshold=threshold)
        rows.append(
            {
                "split": split_name,
                "cells": int(valid.sum()),
                "positive_ratio": float(y_true[valid].mean()),
                "accuracy": metrics["accuracy"],
                "probability_mae": float(np.mean(np.abs(prob[valid].to_numpy() - y_true[valid].to_numpy()))),
                "tn": metrics["tn"],
                "fp": metrics["fp"],
                "fn": metrics["fn"],
                "tp": metrics["tp"],
            }
        )
    return pd.DataFrame(rows)


def build_image_metric_summary(image_df: pd.DataFrame) -> pd.DataFrame:
    """建立 image 層級的整體與 split 指標。

    Args:
        image_df: image-level prediction table。

    Returns:
        指標摘要表。
    """
    rows: list[dict[str, Any]] = []
    for split_name, local in [("overall", image_df), *list(image_df.groupby("split", sort=True))]:
        if local.empty:
            continue
        valid = local.dropna(subset=["true_ratio", "pred_ratio"]).copy()
        if valid.empty:
            continue
        metrics = evaluate_image_predictions(valid, valid["pred_ratio"])
        rows.append(
            {
                "split": split_name,
                "images": int(len(valid)),
                "cells": int(pd.to_numeric(valid["cell_count"], errors="coerce").fillna(0).sum()),
                "true_ratio_mean": float(pd.to_numeric(valid["true_ratio"], errors="coerce").mean()),
                "pred_ratio_mean": float(pd.to_numeric(valid["pred_ratio"], errors="coerce").mean()),
                "image_mae": metrics["image_mae"],
                "image_bias": metrics["image_bias"],
                "image_rmse": metrics["image_rmse"],
                "within_10pp": metrics["within_10pp"],
            }
        )
    return pd.DataFrame(rows)


def write_feature_sheet(
    workbook: Any,
    sheet_name: str,
    frame: pd.DataFrame,
    id_columns: Sequence[str],
    feature_columns: Sequence[str],
) -> None:
    """輸出單一特徵群工作表。

    Args:
        workbook: openpyxl workbook。
        sheet_name: 工作表名稱。
        frame: 要輸出的資料表。
        id_columns: 識別欄位。
        feature_columns: 特徵欄位。
    """
    columns = existing_ordered_columns(frame, [*id_columns, *feature_columns])
    write_dataframe_sheet(workbook, sheet_name, frame.loc[:, columns])


def aggregate_cell_features_to_image(
    cell_df: pd.DataFrame,
    image_df: pd.DataFrame,
    feature_columns: Sequence[str],
) -> pd.DataFrame:
    """將 cell-level 特徵彙整成 image-level 平均與中位數。

    Args:
        cell_df: cell-level prediction table。
        image_df: image-level prediction table。
        feature_columns: 要彙整的特徵欄位。

    Returns:
        image-level 特徵摘要表。
    """
    columns = existing_ordered_columns(cell_df, feature_columns)
    id_columns = existing_ordered_columns(image_df, ["split", "source_folder", "image_key", "image_name", "passage"])
    id_frame = image_df.loc[:, id_columns].drop_duplicates("image_key")
    if not columns:
        return id_frame

    numeric = cell_df.loc[:, ["image_key", *columns]].copy()
    for column in columns:
        numeric[column] = pd.to_numeric(numeric[column], errors="coerce")
    stats = numeric.groupby("image_key", as_index=True)[columns].agg(["mean", "median"])
    stats.columns = [f"{column}_{stat_name}" for column, stat_name in stats.columns]
    return id_frame.merge(stats.reset_index(), on="image_key", how="left")


def excel_cell_value(value: Any) -> Any:
    """將 pandas / numpy 值轉成 openpyxl 可寫入的值。"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        return value.item()
    return value


def iter_excel_rows(frame: pd.DataFrame):
    """逐列產生可寫入 Excel 的資料。"""
    yield list(frame.columns)
    for row in frame.itertuples(index=False, name=None):
        yield [excel_cell_value(value) for value in row]


def write_dataframe_sheet(workbook: Any, sheet_name: str, frame: pd.DataFrame) -> None:
    """以串流方式輸出 DataFrame 工作表。

    Args:
        workbook: openpyxl workbook。
        sheet_name: 工作表名稱。
        frame: 要輸出的資料表。
    """
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.freeze_panes = "A2"
    for row in iter_excel_rows(frame):
        worksheet.append(row)


def write_prediction_summary_sheet(
    workbook: Any,
    metrics_df: pd.DataFrame,
    predictions_df: pd.DataFrame,
) -> None:
    """輸出含指標與 ground truth / prediction 的最後一張工作表。

    Args:
        workbook: openpyxl workbook。
        metrics_df: 指標摘要表。
        predictions_df: ground truth / prediction 表。
    """
    worksheet = workbook.create_sheet(title=PREDICTION_SUMMARY_SHEET)
    worksheet.append(["Metrics"])
    for row in iter_excel_rows(metrics_df):
        worksheet.append(row)
    worksheet.append([])
    worksheet.append(["Predictions"])
    for row in iter_excel_rows(predictions_df):
        worksheet.append(row)


def export_prediction_workbooks(
    output_dir: Path,
    cell_df: pd.DataFrame,
    image_df: pd.DataFrame,
    feature_groups: dict[str, list[str]],
    threshold: float = 0.5,
) -> None:
    """輸出整理版 single-cell 與 image-level prediction Excel。

    Args:
        output_dir: 訓練輸出資料夾。
        cell_df: cell-level prediction table。
        image_df: image-level prediction table。
        feature_groups: 特徵群對照。
        threshold: 將 `cell_prob` 轉成陽性預測的門檻。
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_groups = prediction_workbook_feature_groups(feature_groups)
    cell_df = append_prediction_columns(cell_df, threshold=threshold)
    image_df = append_image_error_columns(image_df)

    cell_id_columns = ["split", "source_folder", "Image", "Cell_ID", "image_key", "passage"]
    cell_summary_columns = [
        "split",
        "source_folder",
        "Image",
        "Cell_ID",
        "image_key",
        "passage",
        "ki67_label",
        "cell_prob",
        "cell_pred",
        "cell_correct",
        "texture_ki67_evidence",
        "intensity_distribution_ki67_evidence",
        "halo_rounding_ki67_evidence",
    ]
    image_summary_columns = [
        "split",
        "source_folder",
        "image_name",
        "image_key",
        "passage",
        "cell_count",
        "true_ratio",
        "pred_ratio",
        "signed_error",
        "abs_error",
        "mean_cell_prob",
        "median_cell_prob",
        "p90_cell_prob",
        "frac_prob_gt_0_3",
        "frac_prob_gt_0_5",
        "frac_prob_gt_0_7",
        "texture_ki67_evidence_mean",
        "texture_ki67_evidence_median",
        "intensity_distribution_ki67_evidence_mean",
        "intensity_distribution_ki67_evidence_median",
        "halo_rounding_ki67_evidence_mean",
        "halo_rounding_ki67_evidence_median",
    ]

    from openpyxl import Workbook

    cell_workbook = output_dir / "ki67_training_single_cell_predictions.xlsx"
    cell_book = Workbook(write_only=True)
    used_names: set[str] = set()
    for group_name, columns in workbook_groups:
        write_feature_sheet(
            cell_book,
            excel_sheet_name(group_name, used_names),
            cell_df,
            cell_id_columns,
            columns,
        )
    write_prediction_summary_sheet(
        cell_book,
        build_cell_metric_summary(cell_df, threshold=threshold),
        cell_df.loc[:, existing_ordered_columns(cell_df, cell_summary_columns)],
    )
    cell_book.save(cell_workbook)

    image_workbook = output_dir / "ki67_training_image_predictions.xlsx"
    image_book = Workbook(write_only=True)
    used_names = set()
    for group_name, columns in workbook_groups:
        image_feature_table = aggregate_cell_features_to_image(cell_df, image_df, columns)
        image_id_columns = ["split", "source_folder", "image_key", "image_name", "passage"]
        write_feature_sheet(
            image_book,
            excel_sheet_name(group_name, used_names),
            image_feature_table,
            image_id_columns,
            [column for column in image_feature_table.columns if column not in set(image_id_columns)],
        )
    write_prediction_summary_sheet(
        image_book,
        build_image_metric_summary(image_df),
        image_df.loc[:, existing_ordered_columns(image_df, image_summary_columns)],
    )
    image_book.save(image_workbook)


def parse_args() -> argparse.Namespace:
    """解析命令列參數。"""
    parser = argparse.ArgumentParser(description="Train the new Ki67 Stage 1 / Stage 2 pipeline.")
    parser.add_argument("--results-dir", type=Path, default=DEFAULT_RESULTS_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_TRAIN_OUTPUT_DIR)
    parser.add_argument("--summary-html", type=Path, default=DEFAULT_SUMMARY_HTML)
    parser.add_argument("--selected-top-k", type=int, default=40)
    parser.add_argument("--formal-config", default=FORMAL_CONFIG_KEY)
    parser.add_argument("--cv-splits", type=int, default=5)
    parser.add_argument("--skip-cnn", action="store_true")
    parser.add_argument("--cnn-pretrained", action="store_true")
    parser.add_argument("--keep-extreme-image-ratios", action="store_true")
    parser.add_argument("--save-prediction-table", action="store_true")
    return parser.parse_args()


def build_stage1_matrix(
    frame: pd.DataFrame,
    config: ExperimentConfig,
    evidence_scores: pd.DataFrame,
    image_embeddings: pd.DataFrame,
    selected_transformer: Any,
    all_parameter_transformer: Any,
) -> pd.DataFrame:
    """依實驗設定組合 Stage 1 input matrix。"""
    parts: list[pd.DataFrame] = []
    if config.include_feature_scores:
        score_columns = feature_score_columns_for_config(config, evidence_scores)
        if score_columns:
            parts.append(evidence_scores.reindex(columns=score_columns).reset_index(drop=True))
    if config.include_cnn_embedding:
        parts.append(attach_image_embeddings(frame, image_embeddings).reset_index(drop=True))
    if config.parameter_mode == "selected":
        parts.append(transform_table_features(frame, selected_transformer).reset_index(drop=True))
    elif config.parameter_mode == "all":
        parts.append(transform_table_features(frame, all_parameter_transformer).reset_index(drop=True))

    if not parts:
        raise ValueError(f"實驗 {config.key} 沒有任何 Stage 1 輸入欄位。")
    matrix = pd.concat(parts, axis=1)
    matrix = matrix.loc[:, ~matrix.columns.duplicated()].copy()
    return matrix.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def attach_scores_and_probability(
    frame: pd.DataFrame,
    probability: np.ndarray,
    evidence_scores: pd.DataFrame,
) -> pd.DataFrame:
    """建立含 evidence score 與 cell probability 的 cell-level table。"""
    scored = frame.reset_index(drop=True).copy()
    for column in evidence_scores.columns:
        scored[column] = evidence_scores.reset_index(drop=True)[column].to_numpy(dtype=np.float64)
    scored["cell_prob"] = np.asarray(probability, dtype=np.float64)
    return scored


def evaluate_config(
    config: ExperimentConfig,
    train_df: pd.DataFrame,
    valid_df: pd.DataFrame,
    test_df: pd.DataFrame,
    train_evidence: pd.DataFrame,
    valid_evidence: pd.DataFrame,
    test_evidence: pd.DataFrame,
    image_embeddings: pd.DataFrame,
    selected_transformer: Any,
    all_parameter_transformer: Any,
    cv_splits: int,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """訓練並評估單一實驗設定。

    Args:
        config: 實驗設定。
        train_df: Train split cell-level table。
        valid_df: Valid split cell-level table。
        test_df: Test split cell-level table。
        train_evidence: Train OOF feature-group evidence scores。
        valid_evidence: Valid feature-group evidence scores。
        test_evidence: Test feature-group evidence scores。
        image_embeddings: image_key indexed CNN embedding table。
        selected_transformer: Selected parameters 前處理器。
        all_parameter_transformer: All parameters 前處理器。
        cv_splits: Stage 1 GroupKFold 切分數。

    Returns:
        tuple[dict[str, Any], dict[str, Any]]: metric row 與可供部署的模型 artifacts。
    """
    config_score_columns = (
        feature_score_columns_for_config(config, train_evidence)
        if config.include_feature_scores
        else []
    )
    train_evidence_config = train_evidence.reindex(columns=config_score_columns)
    valid_evidence_config = valid_evidence.reindex(columns=config_score_columns)
    test_evidence_config = test_evidence.reindex(columns=config_score_columns)

    x_train = build_stage1_matrix(
        train_df,
        config,
        train_evidence,
        image_embeddings,
        selected_transformer,
        all_parameter_transformer,
    )
    x_valid = build_stage1_matrix(
        valid_df,
        config,
        valid_evidence,
        image_embeddings,
        selected_transformer,
        all_parameter_transformer,
    ).reindex(columns=x_train.columns, fill_value=0.0)
    x_test = build_stage1_matrix(
        test_df,
        config,
        test_evidence,
        image_embeddings,
        selected_transformer,
        all_parameter_transformer,
    ).reindex(columns=x_train.columns, fill_value=0.0)

    stage1_model, train_prob = fit_stage1_classifier_with_oof(
        x_train,
        train_df["ki67_label"],
        train_df["image_key"].astype(str),
        random_state=RANDOM_STATE,
        cv_splits=cv_splits,
    )
    valid_prob = predict_positive_probability(stage1_model, x_valid)
    test_prob = predict_positive_probability(stage1_model, x_test)

    train_cell = attach_scores_and_probability(train_df, train_prob, train_evidence_config)
    valid_cell = attach_scores_and_probability(valid_df, valid_prob, valid_evidence_config)
    test_cell = attach_scores_and_probability(test_df, test_prob, test_evidence_config)
    evidence_columns = list(train_evidence_config.columns)

    train_image = aggregate_to_image_features(train_cell, "cell_prob", evidence_columns)
    valid_image = aggregate_to_image_features(valid_cell, "cell_prob", evidence_columns)
    test_image = aggregate_to_image_features(test_cell, "cell_prob", evidence_columns)

    stage2_model, stage2_columns = fit_stage2_ratio_model(train_image, config.stage2_mode)
    train_image["pred_ratio"] = predict_stage2_ratio(train_image, stage2_model, config.stage2_mode, stage2_columns)
    valid_image["pred_ratio"] = predict_stage2_ratio(valid_image, stage2_model, config.stage2_mode, stage2_columns)
    test_image["pred_ratio"] = predict_stage2_ratio(test_image, stage2_model, config.stage2_mode, stage2_columns)

    train_cell_metrics = evaluate_cell_predictions(train_df["ki67_label"], train_prob)
    valid_cell_metrics = evaluate_cell_predictions(valid_df["ki67_label"], valid_prob)
    test_cell_metrics = evaluate_cell_predictions(test_df["ki67_label"], test_prob)
    train_image_metrics = evaluate_image_predictions(train_image, train_image["pred_ratio"])
    valid_image_metrics = evaluate_image_predictions(valid_image, valid_image["pred_ratio"])
    test_image_metrics = evaluate_image_predictions(test_image, test_image["pred_ratio"])

    row = {
        "config_key": config.key,
        "display_name": config.display_name,
        "stage1_input": config.stage1_input_name,
        "stage2_input": config.stage2_name,
        "train_cell_accuracy": train_cell_metrics["accuracy"],
        "valid_cell_accuracy": valid_cell_metrics["accuracy"],
        "test_cell_accuracy": test_cell_metrics["accuracy"],
        "train_image_mae": train_image_metrics["image_mae"],
        "valid_image_mae": valid_image_metrics["image_mae"],
        "test_image_mae": test_image_metrics["image_mae"],
        "train_image_bias": train_image_metrics["image_bias"],
        "valid_image_bias": valid_image_metrics["image_bias"],
        "test_image_bias": test_image_metrics["image_bias"],
        "test_tn": test_cell_metrics["tn"],
        "test_fp": test_cell_metrics["fp"],
        "test_fn": test_cell_metrics["fn"],
        "test_tp": test_cell_metrics["tp"],
    }
    artifacts = {
        "config": asdict(config),
        "stage1_model": stage1_model,
        "stage1_feature_columns": list(x_train.columns),
        "stage2_model": stage2_model,
        "stage2_feature_columns": list(stage2_columns),
        "train_cell_predictions": train_cell,
        "valid_cell_predictions": valid_cell,
        "test_cell_predictions": test_cell,
        "train_image_predictions": train_image,
        "valid_image_predictions": valid_image,
        "test_image_predictions": test_image,
    }
    return row, artifacts


def prepare_training_data(args: argparse.Namespace) -> tuple[dict[str, pd.DataFrame], list[str]]:
    """讀取與切分新版訓練資料。"""
    csv_files = find_cleaned_csv_files(
        results_dir=args.results_dir,
        excluded_source_folders=DEFAULT_EXCLUDED_SOURCE_FOLDERS,
    )
    dataset = load_ki67_dataset(csv_files, require_label=True)
    if not args.keep_extreme_image_ratios:
        dataset = filter_extreme_image_ratios(dataset)
    splits = split_images_within_source_folder(dataset, random_state=RANDOM_STATE)
    feature_columns = detect_numeric_feature_columns(dataset)
    return splits, feature_columns


def run_training(args: argparse.Namespace) -> dict[str, Any]:
    """執行完整新版訓練流程。"""
    output_dir = args.output_dir
    model_dir = output_dir / "model"
    reset_directory(output_dir)
    model_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    splits, feature_columns = prepare_training_data(args)
    train_df = splits["train"]
    valid_df = splits["valid"]
    test_df = splits["test"]
    if len(train_df) == 0 or len(valid_df) == 0 or len(test_df) == 0:
        raise ValueError("Train / valid / test 任一切分為空，請檢查 source folder 設定。")

    feature_groups = build_feature_groups(feature_columns)
    feature_group_models, train_evidence = fit_feature_group_models(
        train_df,
        feature_groups,
        BASE_FEATURE_GROUPS,
        cv_splits=args.cv_splits,
        random_state=RANDOM_STATE,
        group_column="image_key",
    )
    valid_evidence = predict_feature_group_scores(valid_df, feature_group_models)
    test_evidence = predict_feature_group_scores(test_df, feature_group_models)

    selected_transformer = fit_selected_parameter_transformer(train_df, feature_columns, top_k=args.selected_top_k)
    all_parameter_transformer = fit_table_transformer(train_df, feature_columns)

    all_for_embedding = pd.concat([train_df, valid_df, test_df], ignore_index=True)
    if args.skip_cnn:
        image_embeddings = pd.DataFrame(index=[])
        cnn_meta = {"status": "skipped"}
    else:
        image_embeddings, cnn_meta = extract_image_embeddings(
            all_for_embedding,
            pretrained=bool(args.cnn_pretrained),
            random_state=RANDOM_STATE,
        )
    if image_embeddings.empty:
        print("[警告] CNN embedding 為空；包含 CNN 的實驗會使用空矩陣並可能失敗。")

    configs = default_experiment_configs()
    metrics: list[dict[str, Any]] = []
    artifact_by_key: dict[str, dict[str, Any]] = {}
    for config in configs:
        if config.include_cnn_embedding and image_embeddings.empty:
            print(f"[略過] {config.display_name}: 沒有 CNN embedding。")
            continue
        row, artifacts = evaluate_config(
            config,
            train_df,
            valid_df,
            test_df,
            train_evidence,
            valid_evidence,
            test_evidence,
            image_embeddings,
            selected_transformer,
            all_parameter_transformer,
            cv_splits=args.cv_splits,
        )
        metrics.append(row)
        artifact_by_key[config.key] = artifacts

    if not metrics:
        raise RuntimeError("沒有任何實驗成功完成。")
    metrics_df = pd.DataFrame(metrics).sort_values("test_image_mae").reset_index(drop=True)
    formal_key = args.formal_config if args.formal_config in artifact_by_key else str(metrics_df.iloc[0]["config_key"])
    best_artifacts = artifact_by_key[formal_key]
    split_df = pd.DataFrame(
        [
            summarize_split(train_df, "train"),
            summarize_split(valid_df, "valid"),
            summarize_split(test_df, "test"),
        ]
    )

    metrics_out = metrics_df.copy()
    metrics_out.insert(0, "item", "experiment_metric")
    split_out = split_df.copy()
    split_out.insert(0, "item", "split_summary")
    training_summary = pd.concat([metrics_out, split_out], ignore_index=True, sort=False)
    training_summary.to_csv(output_dir / "ki67_training_summary.csv", index=False, encoding="utf-8-sig")

    if args.save_prediction_table:
        prediction_tables: list[pd.DataFrame] = []
        cell_prediction_tables: list[pd.DataFrame] = []
        image_prediction_tables: list[pd.DataFrame] = []
        for split_name in ("train", "valid", "test"):
            cell_table = best_artifacts[f"{split_name}_cell_predictions"].copy()
            cell_table.insert(0, "level", "cell")
            cell_table.insert(0, "split", split_name)
            image_table = best_artifacts[f"{split_name}_image_predictions"].copy()
            image_table.insert(0, "level", "image")
            image_table.insert(0, "split", split_name)
            prediction_tables.extend([cell_table, image_table])
            cell_prediction_tables.append(cell_table.drop(columns=["level"]))
            image_prediction_tables.append(image_table.drop(columns=["level"]))
        pd.concat(prediction_tables, ignore_index=True, sort=False).to_csv(
            output_dir / "ki67_training_predictions.csv",
            index=False,
            encoding="utf-8-sig",
        )
        export_prediction_workbooks(
            output_dir=output_dir,
            cell_df=pd.concat(cell_prediction_tables, ignore_index=True, sort=False),
            image_df=pd.concat(image_prediction_tables, ignore_index=True, sort=False),
            feature_groups=feature_groups,
            threshold=0.5,
        )

    model_bundle = {
        "version": "feature_group_evidence_cnn_s1",
        "formal_config_key": formal_key,
        "random_state": RANDOM_STATE,
        "feature_columns": list(feature_columns),
        "feature_group_names": list(BASE_FEATURE_GROUPS),
        "feature_groups": {name: list(cols) for name, cols in feature_groups.items()},
        "feature_group_models": feature_group_models,
        "selected_parameter_transformer": selected_transformer,
        "all_parameter_transformer": all_parameter_transformer,
        "stage1_model": best_artifacts["stage1_model"],
        "stage1_feature_columns": best_artifacts["stage1_feature_columns"],
        "stage2_model": best_artifacts["stage2_model"],
        "stage2_feature_columns": best_artifacts["stage2_feature_columns"],
        "config": best_artifacts["config"],
        "cnn_meta": cnn_meta,
        "cnn_pretrained": bool(args.cnn_pretrained),
        "cell_threshold": 0.5,
    }
    save_model_bundle(model_dir / "ki67_model_bundle.joblib", model_bundle)
    save_json(
        model_dir / "training_meta.json",
        {
            "formal_config_key": formal_key,
            "cnn_meta": cnn_meta,
            "selected_top_k": int(args.selected_top_k),
            "cv_splits": int(args.cv_splits),
            "summary_html": str(args.summary_html),
        },
    )

    render_summary_html(args.summary_html, metrics_df, split_df, formal_key)

    best_row = metrics_df[metrics_df["config_key"] == formal_key].iloc[0]
    print("===== Ki67 新版訓練完成 =====")
    print(f"模型 bundle: {model_dir / 'ki67_model_bundle.joblib'}")
    print(f"summary.html: {args.summary_html}")
    print(
        f"正式組合: {best_row['display_name']} | "
        f"test image MAE={format_pp(best_row['test_image_mae'])} | "
        f"test cell acc={format_percent(best_row['test_cell_accuracy'])}"
    )
    return {
        "metrics": metrics_df,
        "split_summary": split_df,
        "model_bundle": model_bundle,
        "output_dir": output_dir,
    }


def main() -> None:
    """命令列入口。"""
    run_training(parse_args())


if __name__ == "__main__":
    main()
