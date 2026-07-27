"""將合併後的 cleaned CSV 匯出為工程版或生醫版 Excel 工作簿。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

WorkbookProfile = Literal["engineer", "biomedical"]

META_FEATURES = {"Cell_ID", "Image", "cell_status"}
GEOMETRY_BASES = {
    "Area",
    "Perimeter",
    "Convex Perimeter",
    "Circular Diameter",
    "Feret Length",
    "Feret Width",
    "Extent",
    "Major Axis Length",
    "Minor Axis Length",
    "Maximum Radius",
    "Mean Radius",
    "Median Radius",
    "Perimeter/Area Ratio",
    "Solidity",
    "Aspect Ratio",
    "Eccentricity",
    "Roundness",
    "Circularity",
    "Sphericity",
    "Roughness",
    "Compactness",
}
INTENSITY_BASES = {
    "Mean",
    "StdDev",
    "Min",
    "Max",
    "IntDen",
    "RawIntDen",
    "CV",
    "Range",
    "P10",
    "P25",
    "P75",
    "P90",
    "IQR80",
    "Entropy",
    "Skewness",
    "Kurtosis",
}
DERIVED_FEATURES = {
    "Karyoplasmic Ratio",
    "Nuc Cyto Mean Ratio",
    "Nuc Cyto IntDen Ratio",
    "Nuc Cyto RawIntDen Ratio",
    "Nuc Cell IntDen Ratio",
    "Nuc Cyto Entropy Difference",
    "Nuc Cyto CV Difference",
    "Nucleus Centroid Offset",
}
NUCLEOLUS_FEATURES = {
    "Nucleolus Count",
    "Mean Nucleolus Area",
    "Max Nucleolus Area",
}
HALO_FEATURES = {
    "Halo Outer Mean",
    "Halo Outer StdDev",
    "Halo Outer CV",
    "Halo Inner Mean",
    "Halo Inner StdDev",
    "Halo Inner Outer Diff",
    "Halo Angular Variance",
    "Halo Radial Gradient",
    "Halo Width",
    "Edge Sharpness",
}
LOCAL_CROWDING_FEATURES = {
    "Nearest Neighbor Distance",
    "Nearest Neighbor Distance Norm",
    "Local Neighbor Count",
    "Local Density",
    "Neighbour Area Ratio",
}
COLONY_FEATURES = {
    "Image Confluency",
    "Population Area CV",
    "Population Circularity CV",
    "Cluster Size",
    "Cluster Size Norm",
    "Largest Cluster Ratio",
}
POPULATION_FEATURES = LOCAL_CROWDING_FEATURES | COLONY_FEATURES | {"Mitotic Index"}
SHAPE_FEATURES = {
    "Protrusion Count",
    "Mean Convex Defect Depth",
    "Mean Protrusion Length Norm",
    "Max Convex Defect Depth",
    "Fractal Dimension",
    "Boundary Inflection Count",
}
MITOSIS_FEATURES = {
    "Mitotic Index",
    "Mitotic Score",
    "Daughter Pair Flag",
    "Protrusion Retraction Score",
}
DEBRIS_FEATURES = {
    "Debris Count",
    "Debris Area Fraction",
    "Nearest Debris Distance",
    "Debris Mean Area",
    "Debris Density",
}


@dataclass
class FeatureGroup:
    """描述一個輸出工作表及其特徵欄位。"""

    name: str
    description: str
    color: str
    features: list[str]


GROUP_SPECS: dict[WorkbookProfile, list[tuple[str, str, str]]] = {
    "engineer": [
        ("Texture", "GLCM、LBP、Tamura 與 Zernike 紋理特徵。", "8064A2"),
        (
            "Intensity distribution",
            "核、細胞輪廓與全細胞的強度分布及衍生比值。",
            "00A6A6",
        ),
        (
            "Attachment-spreading",
            "細胞鋪展、細胞輪廓幾何、突起與中心偏移。",
            "ED7D31",
        ),
        ("Halo-rounding", "細胞邊緣與 phase-contrast halo 特徵。", "FFC000"),
        ("Local crowding", "單一細胞附近的距離、鄰居數與局部密度。", "70AD47"),
        (
            "Colony-FOV context",
            "影像層級匯合度、群體變異與群集資訊。",
            "548235",
        ),
        ("Mitosis likelihood", "有絲分裂樣的規則式形態指標。", "C00000"),
        ("Nuclear morphology", "細胞核幾何、核質比與核仁候選。", "4472C4"),
        (
            "Debris-culture health",
            "影像背景碎屑與培養健康 proxy。",
            "7F7F7F",
        ),
        ("IDO response", "IDO 螢光原始與背景校正量測。", "A64D79"),
        ("Other", "尚未納入既有分群的新欄位。", "767171"),
    ],
    "biomedical": [
        ("大小與尺度", "細胞／細胞核的面積與直徑。", "4F81BD"),
        ("形狀與延展", "圓形、細長與延展程度。", "2F75B5"),
        ("邊界與不規則", "邊界粗糙、突起、光暈與輪廓複雜度。", "ED7D31"),
        ("核與細胞關係", "核質比、核／細胞強度關係與核仁候選。", "8064A2"),
        ("強度與分布", "影像亮暗、分布寬度與統計形狀。", "00A6A6"),
        ("紋理與異質性", "細胞內部紋理規則性與空間異質性。", "5B5EA6"),
        ("群體與排列", "匯合度、鄰居、密度與群集排列。", "70AD47"),
        (
            "分裂碎屑與健康",
            "分裂樣態與背景碎屑等培養狀態 proxy。",
            "C0504D",
        ),
        (
            "IDO反應",
            "IDO 螢光原始與背景校正量測；不是免疫能力的直接 ground truth。",
            "A64D79",
        ),
        ("其他欄位", "尚未納入既有分群的新欄位。", "767171"),
    ],
}

REGION_LABELS = {
    "nucleus": ("細胞核", "細胞核 ROI"),
    "cell_outline": ("細胞輪廓", "細胞輪廓 ROI（依目前程式實作，包含細胞核）"),
    "whole_cell": ("全細胞", "全細胞 ROI"),
    "none": ("", ""),
}

GEOMETRY_INFO = {
    "Area": ("面積", "ROI 所含像素面積。"),
    "Perimeter": ("周長", "ROI 邊界的總長度。"),
    "Convex Perimeter": ("凸包周長", "包住 ROI 的最小凸多邊形之周長。"),
    "Circular Diameter": ("等效圓直徑", "與 ROI 面積相同之圓的直徑。"),
    "Feret Length": ("最大 Feret 直徑", "所有方向上量得的最大卡尺距離。"),
    "Feret Width": ("最小 Feret 直徑", "所有方向上量得的最小卡尺距離。"),
    "Extent": ("外接矩形填充率", "面積／外接矩形面積。"),
    "Major Axis Length": ("主軸長度", "具有相同二階矩之橢圓的長軸長度。"),
    "Minor Axis Length": ("次軸長度", "具有相同二階矩之橢圓的短軸長度。"),
    "Maximum Radius": ("最大半徑", "ROI 像素到最近背景像素距離的最大值。"),
    "Mean Radius": ("平均半徑", "ROI 像素到最近背景像素距離的平均值。"),
    "Median Radius": ("中位半徑", "ROI 像素到最近背景像素距離的中位數。"),
    "Perimeter/Area Ratio": ("周長／面積比", "周長除以面積。"),
    "Solidity": ("實心度", "面積／凸包面積。"),
    "Aspect Ratio": ("長寬比", "主軸長度除以次軸長度；越大代表越細長。"),
    "Eccentricity": ("離心率", "越接近 1 代表形狀越細長。"),
    "Roundness": ("圓整度", "依 ImageJ roundness 或主軸與面積估算。"),
    "Circularity": ("專案圓形度", "反映 ROI 接近圓形的程度；需留意專案公式。"),
    "Sphericity": ("形狀因子", "4π×Area／Perimeter²；越接近 1 越像圓。"),
    "Roughness": ("邊界粗糙度", "凸包周長／實際周長，反映邊界凹凸。"),
    "Compactness": (
        "緊密度",
        "CellProfiler 相容定義：2π×ROI 像素至形心的平均平方距離／面積。",
    ),
}

INTENSITY_INFO = {
    "Mean": ("平均強度", "ROI 內像素強度的平均值。"),
    "StdDev": ("強度標準差", "ROI 內像素強度的標準差。"),
    "Min": ("最小強度", "ROI 內最小像素強度。"),
    "Max": ("最大強度", "ROI 內最大像素強度。"),
    "IntDen": ("積分強度", "ROI 的 integrated density。"),
    "RawIntDen": ("原始積分強度", "ROI 內原始像素值總和。"),
    "CV": ("強度變異係數", "強度標準差／平均強度。"),
    "Range": ("強度範圍", "最大強度－最小強度。"),
    "P10": ("第 10 百分位強度", "ROI 強度分布的第 10 百分位。"),
    "P25": ("第 25 百分位強度", "ROI 強度分布的第 25 百分位。"),
    "P75": ("第 75 百分位強度", "ROI 強度分布的第 75 百分位。"),
    "P90": ("第 90 百分位強度", "ROI 強度分布的第 90 百分位。"),
    "IQR80": ("中央 80% 強度範圍", "P90－P10，降低極端值影響。"),
    "Entropy": ("強度熵", "越高代表強度分布越分散。"),
    "Skewness": ("強度偏態", "強度分布左右不對稱的程度。"),
    "Kurtosis": ("強度峰度", "強度分布尖峭與尾端厚度的指標。"),
}

EXACT_INFO = {
    "Karyoplasmic Ratio": ("核質比", "細胞核面積／（細胞輪廓面積－細胞核面積）。"),
    "Nuc Cyto Mean Ratio": ("核／細胞輪廓平均強度比", "Mean_nuc／Mean_cyto。"),
    "Nuc Cyto IntDen Ratio": ("核／細胞輪廓積分強度比", "IntDen_nuc／IntDen_cyto。"),
    "Nuc Cyto RawIntDen Ratio": (
        "核／細胞輪廓原始積分強度比",
        "RawIntDen_nuc／RawIntDen_cyto。",
    ),
    "Nuc Cell IntDen Ratio": (
        "核／全細胞積分強度比",
        "IntDen_nuc／Whole Cell IntDen。",
    ),
    "Nuc Cyto Entropy Difference": ("核與細胞輪廓熵差", "Entropy_nuc－Entropy_cyto。"),
    "Nuc Cyto CV Difference": ("核與細胞輪廓 CV 差", "CV_nuc－CV_cyto。"),
    "Nucleus Centroid Offset": (
        "細胞核中心偏移",
        "核中心與全細胞中心距離，再以等效細胞尺度正規化。",
    ),
    "Nucleolus Count": ("核仁候選數", "細胞核內明亮局部極大區域的候選數量。"),
    "Mean Nucleolus Area": ("平均核仁候選面積", "核仁候選連通區面積的平均值。"),
    "Max Nucleolus Area": ("最大核仁候選面積", "最大核仁候選連通區面積。"),
    "Halo Outer Mean": ("外圈光暈平均強度", "細胞外側環帶的平均強度。"),
    "Halo Outer StdDev": ("外圈光暈強度標準差", "細胞外側環帶強度的標準差。"),
    "Halo Outer CV": ("外圈光暈強度 CV", "Halo Outer StdDev／Halo Outer Mean。"),
    "Halo Inner Mean": ("內圈邊緣平均強度", "細胞內側邊緣環帶的平均強度。"),
    "Halo Inner StdDev": ("內圈邊緣強度標準差", "細胞內側環帶強度的標準差。"),
    "Halo Inner Outer Diff": ("內外圈光暈強度差", "內圈與外圈平均強度的絕對差。"),
    "Halo Angular Variance": ("光暈角向變異", "外圈不同角度扇區平均強度的變異。"),
    "Halo Radial Gradient": ("光暈徑向梯度", "由細胞邊界向外的強度變化斜率。"),
    "Halo Width": ("光暈寬度", "由細胞邊界向外回到背景範圍的估計距離。"),
    "Edge Sharpness": ("邊緣銳利度", "細胞邊界環帶上的邊緣反應平均值。"),
    "Image Confluency": ("影像匯合度", "影像中所有細胞面積總和／影像面積。"),
    "Population Area CV": ("群體面積 CV", "同一影像內細胞面積的變異係數。"),
    "Population Circularity CV": ("群體圓形度 CV", "同一影像內細胞圓形度的變異係數。"),
    "Nearest Neighbor Distance": ("最近鄰距離", "該細胞中心到最近其他細胞中心的距離。"),
    "Nearest Neighbor Distance Norm": (
        "正規化最近鄰距離",
        "最近鄰距離／中位等效直徑。",
    ),
    "Local Neighbor Count": ("局部鄰居數", "指定搜尋半徑內的鄰居數。"),
    "Local Density": ("局部密度", "局部鄰居數／搜尋圓面積。"),
    "Neighbour Area Ratio": ("鄰近細胞面積比", "該細胞面積／最近鄰居平均面積。"),
    "Cluster Size": ("群集大小", "該細胞所屬群集的細胞數。"),
    "Cluster Size Norm": ("正規化群集大小", "Cluster Size／該影像細胞總數。"),
    "Largest Cluster Ratio": ("最大群集比例", "最大群集細胞數／影像細胞總數。"),
    "Mitotic Index": ("有絲分裂樣比例", "同一影像內有絲分裂樣細胞的形態 proxy。"),
    "Protrusion Count": ("突起數", "超過深度門檻的凸包缺陷數量。"),
    "Mean Convex Defect Depth": ("平均凸包缺陷深度", "有效凸包缺陷深度的平均值。"),
    "Mean Protrusion Length Norm": (
        "正規化平均突起長度",
        "平均凸包缺陷深度／細胞等效直徑。",
    ),
    "Max Convex Defect Depth": ("最大凸包缺陷深度", "有效凸包缺陷的最大深度。"),
    "Fractal Dimension": ("邊界分形維度", "以 box-counting 估計細胞邊界複雜度。"),
    "Boundary Inflection Count": ("邊界反曲點數", "細胞輪廓曲率符號改變的次數。"),
    "Debris Count": ("碎屑數", "背景區域中符合大小條件的碎屑候選數量。"),
    "Debris Area Fraction": ("碎屑面積比例", "碎屑總面積／可分析背景面積。"),
    "Nearest Debris Distance": ("最近碎屑距離", "細胞中心到最近碎屑中心的距離。"),
    "Debris Mean Area": ("平均碎屑面積", "碎屑候選的平均面積。"),
    "Debris Density": ("碎屑密度", "碎屑數／可分析背景面積。"),
    "Mitotic Score": ("有絲分裂樣分數", "結合多項形態特徵的有絲分裂樣分數。"),
    "Daughter Pair Flag": ("子細胞對標記", "近鄰且面積相近時標為 1。"),
    "Protrusion Retraction Score": (
        "突起回縮分數",
        "結合圓形度、面積與突起的形態分數。",
    ),
    "IDO_BackgroundMean": ("IDO 背景平均強度", "細胞外背景區域的平均螢光強度。"),
    "IDO_MeanIntensity": ("IDO 細胞平均強度", "細胞質區域的 IDO 平均強度，未扣背景。"),
    "IDO_MeanIntensity_BgSub": ("IDO 背景校正平均強度", "細胞平均強度－背景平均強度。"),
    "IDO_IntDen": ("IDO 細胞積分強度", "細胞質區域的 IDO 螢光總量，未扣背景。"),
    "IDO_IntDen_BgSub": ("IDO 背景校正積分強度", "扣除背景後的 IDO 積分強度。"),
}

PAPER_FEATURES = [
    (
        "面積",
        "Area",
        "ROI 內像素面積。",
        "已建置",
        "Area_nuc；Area_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "緊密度",
        "Compactness",
        "描述形狀緊密程度。",
        "已建置",
        "Compactness_nuc；Compactness_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "離心率",
        "Eccentricity",
        "橢圓離心程度。",
        "已建置",
        "Eccentricity_nuc；Eccentricity_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "外接矩形填充率",
        "Extent",
        "Area／bounding-box area。",
        "已建置",
        "Extent_nuc；Extent_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "形狀因子",
        "FormFactor",
        "4π×Area／Perimeter²。",
        "可由現有欄位對應",
        "Sphericity_nuc；Sphericity_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "主軸長度",
        "MajorAxisLength",
        "等二階矩橢圓的長軸長度。",
        "已建置",
        "Major Axis Length_nuc；Major Axis Length_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "最大 Feret 直徑",
        "MaximumFeretDiameter",
        "所有方向中的最大卡尺距離。",
        "已建置",
        "Feret Length_nuc；Feret Length_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "次軸長度",
        "MinorAxisLength",
        "等二階矩橢圓的短軸長度。",
        "已建置",
        "Minor Axis Length_nuc；Minor Axis Length_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "最小 Feret 直徑",
        "MinimumFeretDiameter",
        "所有方向中的最小卡尺距離。",
        "已建置",
        "Feret Width_nuc；Feret Width_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "最大半徑",
        "MaximumRadius",
        "ROI 像素到最近背景像素距離的最大值。",
        "已建置",
        "Maximum Radius_nuc；Maximum Radius_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "平均半徑",
        "MeanRadius",
        "ROI 像素到最近背景像素距離的平均值。",
        "已建置",
        "Mean Radius_nuc；Mean Radius_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "中位半徑",
        "MedianRadius",
        "ROI 像素到最近背景像素距離的中位數。",
        "已建置",
        "Median Radius_nuc；Median Radius_cyto",
        "_paper_shape_measurements()",
    ),
    (
        "長寬比",
        "AspectRatio",
        "主軸長度／次軸長度。",
        "已建置",
        "Aspect Ratio_nuc；Aspect Ratio_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "周長／面積比",
        "Perimeter/Area Ratio",
        "Perimeter／Area。",
        "已建置",
        "Perimeter/Area Ratio_nuc；Perimeter/Area Ratio_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "周長",
        "Perimeter",
        "ROI 邊界總長度。",
        "已建置",
        "Perimeter_nuc；Perimeter_cyto",
        "_geometry_from_measurements()",
    ),
    (
        "實心度",
        "Solidity",
        "Area／convex-hull area。",
        "已建置",
        "Solidity_nuc；Solidity_cyto",
        "_paper_shape_measurements()",
    ),
]


def workbook_output_path(cleaned_csv: Path, profile: WorkbookProfile) -> Path:
    """依輸出版本取得工作簿路徑。

    Args:
        cleaned_csv: 檔名必須以 ``_cleaned.csv`` 結尾的來源檔。
        profile: ``engineer`` 或 ``biomedical``。

    Returns:
        對應版本的 XLSX 路徑。

    Raises:
        ValueError: 來源不是 cleaned CSV，或版本名稱不支援。
    """
    cleaned_csv = Path(cleaned_csv)
    if not cleaned_csv.name.endswith("_cleaned.csv"):
        raise ValueError("新版工作簿只接受檔名以 _cleaned.csv 結尾的來源。")
    if profile == "biomedical":
        return cleaned_csv.with_suffix(".xlsx")
    if profile == "engineer":
        return cleaned_csv.with_name(f"{cleaned_csv.stem}_se.xlsx")
    raise ValueError(f"不支援的工作簿版本：{profile}")


def _parse_feature_name(feature: str) -> tuple[str, str]:
    if feature.startswith("Whole Cell "):
        return feature[len("Whole Cell ") :], "whole_cell"
    if feature.endswith("_nuc"):
        return feature[:-4], "nucleus"
    if feature.endswith("_cyto"):
        return feature[:-5], "cell_outline"
    return feature, "none"


def _is_texture_base(base: str) -> bool:
    return (
        base.startswith("GLCM ")
        or base.startswith("LBP ")
        or base == "Tamura Coarseness"
        or base.startswith("Zernike Moment ")
    )


def _classify_feature(feature: str, profile: WorkbookProfile) -> str | None:
    if feature in META_FEATURES:
        return None
    base, region = _parse_feature_name(feature)
    if feature.startswith("IDO_"):
        return "IDO response" if profile == "engineer" else "IDO反應"

    if profile == "engineer":
        if _is_texture_base(base):
            return "Texture"
        if base in INTENSITY_BASES or (
            feature in DERIVED_FEATURES
            and feature not in {"Nucleus Centroid Offset", "Karyoplasmic Ratio"}
        ):
            return "Intensity distribution"
        if feature in HALO_FEATURES:
            return "Halo-rounding"
        if feature in LOCAL_CROWDING_FEATURES:
            return "Local crowding"
        if feature in COLONY_FEATURES:
            return "Colony-FOV context"
        if feature in MITOSIS_FEATURES:
            return "Mitosis likelihood"
        if feature in DEBRIS_FEATURES:
            return "Debris-culture health"
        if (
            region == "nucleus"
            or feature == "Karyoplasmic Ratio"
            or feature in NUCLEOLUS_FEATURES
        ):
            return "Nuclear morphology"
        if (
            region == "cell_outline"
            or feature in SHAPE_FEATURES
            or feature == "Nucleus Centroid Offset"
        ):
            return "Attachment-spreading"
        return "Other"

    if feature in DERIVED_FEATURES or feature in NUCLEOLUS_FEATURES:
        return "核與細胞關係"
    if _is_texture_base(base):
        return "紋理與異質性"
    if base in INTENSITY_BASES:
        return "強度與分布"
    if base in {"Area", "Circular Diameter", "Feret Length", "Feret Width"}:
        return "大小與尺度"
    if base in {
        "Aspect Ratio",
        "Eccentricity",
        "Roundness",
        "Circularity",
        "Sphericity",
    }:
        return "形狀與延展"
    if (
        base in {"Perimeter", "Convex Perimeter", "Roughness", "Compactness"}
        or feature in HALO_FEATURES
        or feature in SHAPE_FEATURES
    ):
        return "邊界與不規則"
    if feature in POPULATION_FEATURES:
        return "群體與排列"
    if feature in MITOSIS_FEATURES or feature in DEBRIS_FEATURES:
        return "分裂碎屑與健康"
    return "其他欄位"


def _build_groups(columns: list[str], profile: WorkbookProfile) -> list[FeatureGroup]:
    groups = {
        name: FeatureGroup(name, description, color, [])
        for name, description, color in GROUP_SPECS[profile]
    }
    for feature in columns:
        group_name = _classify_feature(feature, profile)
        if group_name is not None:
            groups[group_name].features.append(feature)
    result = [
        groups[name] for name, _, _ in GROUP_SPECS[profile] if groups[name].features
    ]
    assigned = [feature for group in result for feature in group.features]
    expected = [feature for feature in columns if feature not in META_FEATURES]
    if len(assigned) != len(expected) or set(assigned) != set(expected):
        raise ValueError("特徵分群未完整且唯一地涵蓋 cleaned CSV 欄位。")
    return result


def _texture_info(base: str) -> tuple[str, str]:
    known = {
        "GLCM ASM": ("GLCM 角二階矩", "越高代表紋理越規則。"),
        "GLCM Contrast": ("GLCM 對比度", "局部灰階差異程度。"),
        "GLCM Correlation": ("GLCM 相關性", "相鄰灰階線性關聯程度。"),
        "GLCM Difference Variance": ("GLCM 差異變異", "灰階差值分布的變異程度。"),
        "GLCM Entropy": ("GLCM 熵", "越高代表紋理越複雜。"),
        "GLCM Homogeneity": ("GLCM 同質性", "越高代表紋理較均一。"),
        "LBP Mean": ("LBP 平均值", "局部二值模式編碼的平均值。"),
        "LBP StdDev": ("LBP 標準差", "局部二值模式編碼的標準差。"),
        "LBP Entropy": ("LBP 熵", "LBP 直方圖的 entropy。"),
        "LBP Uniform Ratio": ("LBP 均勻模式比例", "uniform LBP pattern 的像素比例。"),
        "Tamura Coarseness": ("Tamura 粗糙度", "估計紋理顆粒尺度。"),
    }
    if base in known:
        return known[base]
    match = re.match(r"^LBP Hist Bin (\d{2})$", base)
    if match:
        return f"LBP 直方圖第 {match.group(1)} 箱", "16-bin LBP 直方圖中的比例。"
    match = re.match(r"^LBP Uniform R([123]) Hist Bin (\d{2})$", base)
    if match:
        return (
            f"LBP 均勻模式 R{match.group(1)} 第 {match.group(2)} 箱",
            "rotation-invariant uniform LBP 直方圖比例。",
        )
    match = re.match(r"^Zernike Moment (\d{2})$", base)
    if match:
        return f"Zernike 矩第 {match.group(1)} 項", "旋轉不變的 Zernike 描述值。"
    return base, "紋理描述特徵。"


def feature_info(feature: str) -> tuple[str, str]:
    """取得特徵的中文名稱與白話定義。"""
    if feature in EXACT_INFO:
        return EXACT_INFO[feature]
    base, region = _parse_feature_name(feature)
    if base in GEOMETRY_INFO:
        chinese, definition = GEOMETRY_INFO[base]
    elif base in INTENSITY_INFO:
        chinese, definition = INTENSITY_INFO[base]
    elif _is_texture_base(base):
        chinese, definition = _texture_info(base)
    else:
        chinese, definition = base, "專案輸出的形態或影像量測特徵。"
    region_name, region_definition = REGION_LABELS[region]
    prefix = f"{region_name}－" if region_name else ""
    definition_prefix = f"{region_definition}：" if region_definition else ""
    return f"{prefix}{chinese}", f"{definition_prefix}{definition}"


def function_info(feature: str) -> str:
    """取得工程版 README 顯示的主要計算函式。"""
    if feature.startswith("IDO_"):
        return "ido_anal()"
    if feature in DERIVED_FEATURES:
        return "_derived_feature_parameter_values()"
    if feature in NUCLEOLUS_FEATURES:
        return "_nucleolus_feature_values()"
    if feature in HALO_FEATURES:
        return (
            "_halo_feature_parameter_values_python() / _halo_feature_parameter_values()"
        )
    if feature in SHAPE_FEATURES:
        return "_shape_complexity_feature_values()"
    if feature in DEBRIS_FEATURES:
        return "_debris_feature_values_for_cells_python() / _debris_feature_values_for_cells()"
    if feature in MITOSIS_FEATURES:
        return "_mitosis_feature_values() / param_anal()"
    if feature in POPULATION_FEATURES:
        return "param_anal() 群體統計"
    base, _ = _parse_feature_name(feature)
    if base in GEOMETRY_BASES:
        return "_measure_roi_with_python() / _measure_roi_with_imagej()"
    if base in INTENSITY_BASES:
        return "_intensity_feature_parameter_values()"
    if base.startswith("GLCM "):
        return "_texture_feature_parameter_values_python() / _texture_feature_parameter_values()"
    if base.startswith("LBP Uniform R"):
        return "_multiradius_lbp_feature_values_python()"
    if base.startswith("LBP "):
        return (
            "_lbp_feature_parameter_values_python() / _lbp_feature_parameter_values()"
        )
    if base == "Tamura Coarseness":
        return "_tamura_coarseness_feature_value()"
    if base.startswith("Zernike Moment "):
        return "_zernike_feature_values_python()"
    return "param_anal()"


def _excel_value(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color.removeprefix("#"))


def _set_internal_link(
    cell: Cell,
    sheet_name: str,
    target_cell: str = "A1",
    *,
    white_text: bool = False,
) -> None:
    escaped = sheet_name.replace("'", "''")
    cell.hyperlink = f"#'{escaped}'!{target_cell}"
    if white_text:
        cell.font = Font(color="FFFFFF", bold=True, underline="single")
    else:
        cell.style = "Hyperlink"


def _write_feature_sheet(
    sheet: Worksheet,
    frame: pd.DataFrame,
    group: FeatureGroup,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = group.color
    headers = ["Cell_ID", "Cell_type", "Image", *group.features]
    sheet.append(headers)
    for _, record in frame.iterrows():
        sheet.append(
            [
                _excel_value(record.get("Cell_ID")),
                _excel_value(record.get("cell_status")),
                _excel_value(record.get("Image")),
                *[_excel_value(record.get(feature)) for feature in group.features],
            ]
        )

    header_fill = _fill(group.color)
    header_border = Border(bottom=Side(style="medium", color="FFFFFF"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = header_border
    sheet.row_dimensions[1].height = 34

    metadata_fill = _fill("F7F9FC")
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 24
        for cell in row[:3]:
            cell.fill = metadata_fill
            cell.font = Font(color="243447", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for cell in row[3:]:
            cell.font = Font(color="1F2937", size=9)
            if isinstance(cell.value, (int, float)) and not isinstance(
                cell.value, bool
            ):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.0000E+00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = [42, 16, 42, *([22] * len(group.features))]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "D2"


def _write_reserved_sheet(sheet: Worksheet, profile: WorkbookProfile) -> None:
    engineer = profile == "engineer"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "595959"
    if engineer:
        headers = [
            "中文名稱",
            "英文名稱",
            "定義",
            "函式",
            "目前狀態",
            "現有對應欄位",
            "備註",
        ]
    else:
        headers = ["中文名稱", "英文名稱", "定義", "目前狀態", "現有對應欄位", "說明"]
    sheet.append(headers)
    for zh, en, definition, status, mapping, function in PAPER_FEATURES:
        note = (
            "預留欄位；目前沒有填入假值。"
            if "尚未" in status
            else "請在模型前確認與論文公式一致。"
        )
        if engineer:
            sheet.append([zh, en, definition, function, status, mapping, note])
        else:
            sheet.append([zh, en, definition, status, mapping, note])

    for cell in sheet[1]:
        cell.fill = _fill("595959")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = _fill("F2F2F2")
            cell.font = Font(color="333333", size=10)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
    widths = [20, 28, 48, 48, 22, 40, 36] if engineer else [20, 28, 48, 24, 40, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 34
    sheet.freeze_panes = "E2" if engineer else "D2"


def _write_readme(
    sheet: Worksheet,
    source_csv: Path,
    frame: pd.DataFrame,
    groups: list[FeatureGroup],
    profile: WorkbookProfile,
) -> None:
    engineer = profile == "engineer"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "1F4E78"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = f"{source_csv.stem}｜{'工程版' if engineer else '生醫版'}特徵工作簿"
    sheet["A1"].fill = _fill("1F4E78")
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34

    summary = [
        ("項目", "內容"),
        (
            "目的",
            (
                "供工程人員檢查特徵欄位與計算函式。"
                if engineer
                else "供生醫人員依生物意義閱讀形態、影像與 IDO 特徵。"
            ),
        ),
        ("來源 CSV", str(source_csv.resolve())),
        ("資料規模", f"{len(frame)} 筆細胞；{len(frame.columns)} 個來源欄位"),
        (
            "排列方式",
            "各特徵分頁一列一個細胞；前三欄固定為 Cell_ID、Cell_type、Image。",
        ),
        ("數值處理", "來源 CSV 數值與缺值原樣搬移；不做標準化、PCA 或模型預測。"),
    ]
    for row_index, (label, value) in enumerate(summary, start=3):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.merge_cells(
            start_row=row_index, start_column=2, end_row=row_index, end_column=4
        )
        for cell in sheet[row_index][0:4]:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            cell.fill = _fill("D9EAF7" if row_index == 3 else "F8FBFE")
        if row_index == 3:
            sheet.cell(row_index, 1).font = Font(bold=True, color="1F4E78")
            sheet.cell(row_index, 2).font = Font(bold=True, color="1F4E78")
    guide_header_row = 11
    guide_data_row = 12
    for column, value in enumerate(
        ["特徵群", "分群用途", "特徵數", "細部特徵定義"], start=1
    ):
        cell = sheet.cell(guide_header_row, column, value)
        cell.fill = _fill("4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    detail_rows: dict[str, int] = {}
    note_header_row = guide_data_row + len(groups) + 2
    notes = [
        ("重要說明", "內容"),
        ("_nuc", "細胞核 ROI。"),
        ("_cyto", "目前使用細胞輪廓 mask，包含細胞核；不能直接解讀為純細胞質。"),
        ("Whole Cell", "全細胞 ROI 的強度與紋理系列。"),
        ("Cell_type", "由 cell_status 填入，代表輪廓／分割狀態，不是生物學細胞種類。"),
        ("IDO 欄位", "代表 IDO 螢光量測與背景校正值；不能單獨等同免疫抑制能力。"),
        ("Paper16_預留", "列出論文 16 個形態特徵；未建置者不填入假資料。"),
        (
            "模型分群",
            (
                "工程版沿用專案特徵群命名。"
                if engineer
                else "生醫版按生物意義分類，分群只用於閱讀。"
            ),
        ),
    ]
    for offset, (label, value) in enumerate(notes):
        row_index = note_header_row + offset
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.merge_cells(
            start_row=row_index, start_column=2, end_row=row_index, end_column=4
        )
        for cell in sheet[row_index][0:4]:
            cell.fill = _fill("A9D18E" if offset == 0 else "F3F8EF")
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
        if offset == 0:
            sheet.cell(row_index, 1).font = Font(bold=True, color="1F4E78")
            sheet.cell(row_index, 2).font = Font(bold=True, color="1F4E78")

    detail_row = note_header_row + len(notes) + 2
    for group in groups:
        detail_rows[group.name] = detail_row
        sheet.merge_cells(
            start_row=detail_row, start_column=1, end_row=detail_row, end_column=4
        )
        title = sheet.cell(
            detail_row, 1, f"{group.name}（{len(group.features)} 個特徵）"
        )
        title.fill = _fill(group.color)
        title.font = Font(bold=True, color="FFFFFF", size=12)
        title.alignment = Alignment(horizontal="center", vertical="center")
        _set_internal_link(title, group.name, white_text=True)

        header_row = detail_row + 1
        headers = (
            ["中文名稱", "英文名稱", "定義", "函式"]
            if engineer
            else ["中文名稱", "英文名稱", "定義"]
        )
        for column, value in enumerate(headers, start=1):
            cell = sheet.cell(header_row, column, value)
            cell.fill = _fill("D9E2F3")
            cell.font = Font(bold=True, color="1F4E78")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for feature_index, feature in enumerate(group.features, start=header_row + 1):
            chinese, definition = feature_info(feature)
            values = [chinese, feature, definition]
            if engineer:
                values.append(function_info(feature))
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(feature_index, column, value)
                cell.fill = _fill("F8FAFD")
                cell.font = Font(color="243447", size=10)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            sheet.row_dimensions[feature_index].height = 36
        detail_row = header_row + len(group.features) + 2

    for offset, group in enumerate(groups):
        row_index = guide_data_row + offset
        values = [group.name, group.description, len(group.features), "查看細部定義"]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.fill = _fill("F7F9FC")
            cell.alignment = Alignment(
                horizontal="right" if column == 3 else "left",
                vertical="center",
                wrap_text=True,
            )
        _set_internal_link(sheet.cell(row_index, 1), group.name)
        _set_internal_link(
            sheet.cell(row_index, 4), "README", f"A{detail_rows[group.name]}"
        )
        estimated_lines = max(1, (len(group.description) + 21) // 22)
        sheet.row_dimensions[row_index].height = max(28, estimated_lines * 18)

    paper_row = detail_row
    sheet.merge_cells(
        start_row=paper_row, start_column=1, end_row=paper_row, end_column=4
    )
    paper_cell = sheet.cell(paper_row, 1, "論文 16 個形態特徵建置狀態 → Paper16_預留")
    paper_cell.fill = _fill("595959")
    paper_cell.alignment = Alignment(horizontal="center", vertical="center")
    _set_internal_link(paper_cell, "Paper16_預留", white_text=True)

    widths = [25, 38, 76 if engineer else 88, 50 if engineer else 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(3, note_header_row + len(notes)):
        if sheet.row_dimensions[row_index].height is None:
            sheet.row_dimensions[row_index].height = 30
    sheet.freeze_panes = "A2"


def export_cleaned_workbook(
    cleaned_csv: Path,
    profile: WorkbookProfile = "engineer",
) -> Path:
    """將 cleaned CSV 匯出為指定受眾版本的 XLSX。

    Args:
        cleaned_csv: 合併後的 ``*_cleaned.csv``。
        profile: ``engineer`` 為工程版，``biomedical`` 為生醫版。

    Returns:
        已建立的 XLSX 路徑。

    Raises:
        FileNotFoundError: cleaned CSV 不存在。
        ValueError: 檔名、版本或必要欄位不合法。
    """
    cleaned_csv = Path(cleaned_csv)
    output_path = workbook_output_path(cleaned_csv, profile)
    if not cleaned_csv.exists():
        raise FileNotFoundError(f"找不到 cleaned CSV：{cleaned_csv}")

    frame = pd.read_csv(cleaned_csv)
    required = {"Cell_ID", "Image"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"cleaned CSV 缺少必要欄位：{', '.join(missing)}")
    if frame.columns.duplicated().any():
        raise ValueError("cleaned CSV 含有重複欄名，無法安全匯出工作簿。")

    groups = _build_groups(list(frame.columns), profile)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    for group in groups:
        _write_feature_sheet(workbook.create_sheet(group.name), frame, group)
    _write_reserved_sheet(workbook.create_sheet("Paper16_預留"), profile)
    _write_readme(readme, cleaned_csv, frame, groups, profile)
    workbook.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


__all__ = [
    "WorkbookProfile",
    "export_cleaned_workbook",
    "feature_info",
    "function_info",
    "workbook_output_path",
]
