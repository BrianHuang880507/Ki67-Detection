from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

import pandas as pd
from openpyxl import load_workbook

from ki67dtc.workbook_export import (
    export_cleaned_workbook,
    workbook_output_path,
)


class WorkbookExportTest(unittest.TestCase):
    def _create_cleaned_csv(self, root: Path) -> tuple[Path, list[str]]:
        feature_columns = [
            "Area_nuc",
            "Area_cyto",
            "GLCM ASM_nuc",
            "Mean_nuc",
            "Halo Outer Mean",
            "Nearest Neighbor Distance",
            "Image Confluency",
            "Mitotic Score",
            "Debris Count",
            "IDO_MeanIntensity",
            "Ki67_Label",
        ]
        frame = pd.DataFrame(
            [
                {
                    "Cell_ID": "sample_1",
                    "Image": "sample",
                    **{
                        feature: index + 1.25
                        for index, feature in enumerate(feature_columns)
                    },
                    "cell_status": "full_cell",
                },
                {
                    "Cell_ID": "sample_2",
                    "Image": "sample",
                    **{
                        feature: index + 2.5
                        for index, feature in enumerate(feature_columns)
                    },
                    "cell_status": "cyto_cut",
                },
            ]
        )
        csv_path = root / "demo_cleaned.csv"
        frame.to_csv(csv_path, index=False)
        return csv_path, feature_columns

    def test_output_names_follow_profile_contract(self) -> None:
        cleaned = Path("demo_cleaned.csv")

        self.assertEqual(
            workbook_output_path(cleaned, "biomedical").name,
            "demo_cleaned.xlsx",
        )
        self.assertEqual(
            workbook_output_path(cleaned, "engineer").name,
            "demo_cleaned_se.xlsx",
        )

    def test_non_cleaned_csv_is_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "_cleaned.csv"):
            workbook_output_path(Path("sample_final.csv"), "engineer")

    def test_both_profiles_preserve_features_and_requested_layout(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            csv_path, feature_columns = self._create_cleaned_csv(Path(tmp_dir))

            outputs = {
                profile: export_cleaned_workbook(csv_path, profile)
                for profile in ("engineer", "biomedical")
            }

            self.assertEqual(outputs["engineer"].name, "demo_cleaned_se.xlsx")
            self.assertEqual(outputs["biomedical"].name, "demo_cleaned.xlsx")
            for profile, output_path in outputs.items():
                with self.subTest(profile=profile):
                    workbook = load_workbook(output_path, data_only=False)
                    self.assertEqual(workbook.sheetnames[0], "README")
                    self.assertIn("Paper16_預留", workbook.sheetnames)

                    exported_features: list[str] = []
                    for sheet in workbook.worksheets:
                        if sheet.title in {"README", "Paper16_預留"}:
                            continue
                        headers = [cell.value for cell in sheet[1]]
                        self.assertEqual(headers[:3], ["Cell_ID", "Cell_type", "Image"])
                        exported_features.extend(headers[3:])
                        self.assertEqual(sheet["A2"].alignment.horizontal, "left")
                        self.assertEqual(sheet["B2"].value, "full_cell")
                        self.assertEqual(sheet["D2"].alignment.horizontal, "right")
                        self.assertEqual(sheet["D1"].alignment.horizontal, "center")
                        self.assertEqual(sheet["D1"].alignment.vertical, "center")
                        self.assertEqual(len(sheet.tables), 0)

                    self.assertCountEqual(exported_features, feature_columns)
                    self.assertEqual(
                        len(exported_features), len(set(exported_features))
                    )
                    self.assertIsNotNone(workbook["README"]["A12"].hyperlink)
                    workbook.close()


if __name__ == "__main__":
    unittest.main()
